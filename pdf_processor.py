# -*- coding: utf-8 -*-
"""
Karmasangsthan Bank - Loan PDF Merge Engine
=============================================
Borrower List PDF + Loan Balance PDF মার্জ করে একটা Excel ফাইল বানায়।

মূল বৈশিষ্ট্য:
  1) প্রতিটা পাতার শেষ রো (pdfplumber-এর table-detection বাগের কারণে যেটা এক কলামে
     মিশে যায়) স্বয়ংক্রিয়ভাবে মেরামত করা হয়।
  2) Loan Program-এর নাম থেকে prefix স্বয়ংক্রিয়ভাবে তৈরি হয় (override করা যায়)।
  3) দুই PDF (account_code, loan_case_no) দিয়ে মেলানো হয়, এবং Borrower Name
     মিলিয়ে ক্রস-চেক করা হয়।
  4) যেসব রো-তে Father/Spouse/Village/Union নিশ্চিতভাবে আলাদা করা যায়নি
     (কারণ নাম একাধিক লাইনে wrap হয়ে জট পাকিয়ে গেছে), সেগুলো ভুল বসানোর বদলে
     "⚠ ম্যানুয়াল যাচাই দরকার" ফ্ল্যাগ দিয়ে চিহ্নিত হয় -- ব্যাংকের ডেটাতে ভুল অনুমানের
     চেয়ে স্বচ্ছভাবে ফ্ল্যাগ করা নিরাপদ।
"""

import re
import difflib
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# ০. ছোট হেল্পার ফাংশন
# =============================================================================
def clean(v):
    """None/newline সামলে সাফ স্ট্রিং রিটার্ন করে"""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\n", " ")).strip()


def parse_ddmmyyyy(s):
    """dd/mm/yyyy স্ট্রিং থেকে datetime; না পারলে None"""
    s = clean(s)
    try:
        return datetime.strptime(s, "%d/%m/%Y")
    except (ValueError, TypeError):
        return None


def to_num(v):
    try:
        v = clean(v).replace(",", "")
        if v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def normalize_case_key(case_str):
    """merge-key হিসেবে ব্যবহারের জন্য loan_case normalize করে -- '05' এবং '5'
    কে একই key-তে মেলাতে leading zero বাদ দেয়। সংখ্যা না হলে যেমন আছে তেমনই রাখে।"""
    s = clean(case_str)
    if re.match(r"^\d+$", s):
        return str(int(s))
    return s


# =============================================================================
# ১. Loan Program Prefix তৈরি করা
# =============================================================================
def generate_prefix(program_name, max_words_for_initials=4):
    """
    ১ শব্দ হলে -> পুরো শব্দ (Title case), যেমন 'Own' -> 'Own'
    ২-৪ শব্দ হলে -> প্রতিটি শব্দের প্রথম অক্ষর, আপারকেসে -> 'Jubo Rin' -> 'JR'
    ৪+ শব্দ হলে -> প্রথম max_words_for_initials শব্দের প্রথম অক্ষর নিয়ে prefix
    """
    words = [w for w in re.split(r"\s+", program_name.strip()) if w]
    if not words:
        return "NA"
    if len(words) == 1:
        return words[0][:1].upper() + words[0][1:]
    chosen = words[:max_words_for_initials]
    return "".join(w[0].upper() for w in chosen if w[0].isalpha())


def build_prefix_map(program_names, overrides=None):
    """
    program_names: {program_code: name} ডিকশনারি (PDF থেকে অটো-ডিটেক্ট করা, key = '0101' এর মতো
                    Loan Program কোড -- account_code (যেমন '100101') নয়)
    overrides: {program_code: custom_prefix} - ইউজার UI থেকে override দিলে সেটা প্রাধান্য পাবে
    রিটার্ন করে: {program_code: prefix}  (একই prefix একাধিক program_code-এ বসে গেলে সেটা
                  স্বয়ংক্রিয়ভাবে ডিজঅ্যামবিগুয়েট করা হয়)
    """
    overrides = overrides or {}
    result = {}
    for code, name in program_names.items():
        if code in overrides and overrides[code].strip():
            result[code] = overrides[code].strip()
        else:
            result[code] = generate_prefix(name)

    # -- একই prefix দুই ভিন্ন প্রোগ্রামে বসে গেলে (collision) ডিজঅ্যামবিগুয়েট করা --
    seen = {}
    for code in sorted(result.keys()):  # ধারাবাহিক ক্রমে যাতে ফল পুনরুৎপাদনযোগ্য (reproducible) হয়
        base = result[code]
        if base not in seen:
            seen[base] = code
        else:
            name = program_names.get(code, "")
            m = re.search(r"(\d+)\D*$", name)  # যেমন 'Policy-2' -> '2', '4th Phase' -> '4'
            suffix = m.group(1) if m else code
            new_prefix = f"{base}{suffix}"
            # নতুন prefix-ও যদি কারো সাথে মিলে যায়, শেষ ভরসা হিসেবে program_code জুড়ে দেওয়া
            if new_prefix in seen:
                new_prefix = f"{base}{code}"
            result[code] = new_prefix
            seen[new_prefix] = code
    return result


# =============================================================================
# ২. পাতা থেকে হেডার + টেবিল বের করা (Loan Program নামসহ)
# =============================================================================
def get_sections_and_tables(pdf_path, progress_cb=None):
    """
    রিটার্ন করে: [(program_code, program_name, status_type, account_code, table_rows), ...]
    """
    results = []
    program_names = {}  # code -> name (সব পাতা থেকে সংগ্রহ)

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for pidx, page in enumerate(pdf.pages):
            if progress_cb:
                progress_cb(pidx + 1, total)

            text = page.extract_text() or ""
            lines = text.split("\n")

            events = []
            cur_program_code = None
            cur_program_name = None

            for l in lines:
                mprog = re.match(r"Loan Program\s*:?\s*(\d+)\s+(.*)", l.strip())
                if mprog:
                    cur_program_code = mprog.group(1).strip()
                    cur_program_name = mprog.group(2).strip()
                    if cur_program_code not in program_names:
                        program_names[cur_program_code] = cur_program_name
                    continue

                if ("Overdue" in l or "Regular" in l) and "Account Code" in l:
                    mtype = re.search(r"(Overdue|Regular)", l)
                    mcode = re.search(r"\b(\d{5,6})\b", l)
                    if mtype and mcode:
                        events.append((mtype.group(1), mcode.group(1),
                                       cur_program_code, cur_program_name))

            tabs = page.extract_tables()
            n = min(len(events), len(tabs))
            if len(events) != len(tabs) and events:
                # কিছু সেকশনের টেবিল মিসিং হতে পারে -- সতর্কতা হিসেবে রাখি
                pass
            for i in range(n):
                stype, code, prog_code, prog_name = events[i]
                results.append((prog_code, prog_name, stype, code, tabs[i]))

    return results, program_names


# =============================================================================
# ৩. Loan Balance PDF-এর ভাঙা (page-শেষের) রো মেরামত
# =============================================================================
_LOAN_BAL_RE = re.compile(
    r'^(\d+)\s+(\d+(?:\.\d+)?)\s+(\d{4}-\d{4}-\d{5})\s+(.+?)\s+([+\d][\d+,]{7,})\s+'
    r'(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(\d{2}/\d{2}/\d{4})\s+(.+)$'
)


def repair_loan_balance_row(merged_text):
    """একটাই কলামে মিশে যাওয়া loan-balance রো মেরামত করে dict রিটার্ন করে, না পারলে None"""
    s = clean(merged_text)
    m = _LOAN_BAL_RE.match(s)
    if not m:
        return None
    sl, case, slno, name, phone, sdate, dur, edate, rest = m.groups()
    tokens = rest.split()
    cls = ""
    if tokens and re.match(r"^[A-Za-z]+$", tokens[-1]):
        cls = tokens[-1]
        tokens = tokens[:-1]
    if len(tokens) < 3:
        return None
    return {
        "loan_case": case,
        "customer_name": clean(name),
        "phone": phone,
        "overdue_date": sdate,
        "bal_principal": tokens[-3],
        "bal_interest": tokens[-2],
        "bal_total": tokens[-1],
        "repaired": True,
    }


def parse_loan_balance_pdf(pdf_path, progress_cb=None):
    """
    রিটার্ন করে: rows dict {(code, loan_case): {...}}, program_names {code: name}
    """
    sections, program_names = get_sections_and_tables(pdf_path, progress_cb)
    rows = {}
    for prog_code, prog_name, stype, code, table in sections:
        for row in table:
            if not row:
                continue

            # -- ভাঙা (page-শেষের) রো ধরা --
            if row[0] and (len(row) < 2 or row[1] is None) and len(str(row[0]).split()) > 5:
                repaired = repair_loan_balance_row(row[0])
                if repaired:
                    norm_case = normalize_case_key(repaired["loan_case"])
                    key = (code, norm_case)
                    rows[key] = {
                        "account_code": code,
                        "program_code": prog_code,
                        "program_name": prog_name,
                        "status_type": stype,
                        "loan_case": norm_case,
                        "customer_name": repaired["customer_name"],
                        "phone": repaired["phone"],
                        "overdue_date": repaired["overdue_date"],
                        "bal_principal": repaired["bal_principal"],
                        "bal_interest": repaired["bal_interest"],
                        "bal_total": repaired["bal_total"],
                        "needs_review": False,
                    }
                continue

            # -- স্বাভাবিক রো --
            if not row[1] or not re.match(r"^\d+(\.\d+)?$", str(row[1]).strip()):
                continue
            loan_case = normalize_case_key(row[1])
            key = (code, loan_case)

            def g(i):
                return row[i] if i < len(row) else None

            rows[key] = {
                "account_code": code,
                "program_code": prog_code,
                "program_name": prog_name,
                "status_type": stype,
                "loan_case": loan_case,
                "customer_name": clean(g(3)),
                # NOTE: index ৪ (phone) আর ৫ (overdue_date) repair_loan_balance_row()-এর
                # regex-এ পাওয়া কলাম-অর্ডার (serial, case, slno, name, phone, sdate, ...)
                # থেকে অনুমান করা -- আসল balance PDF-এ extract_tables() কী index দিচ্ছে
                # সেটা প্রথম রানের পর print(row) দিয়ে মিলিয়ে নিশ্চিত হয়ে নিন।
                "phone": clean(g(4)),
                "overdue_date": clean(g(7)),
                "bal_principal": clean(g(14)),
                "bal_interest": clean(g(15)),
                "bal_total": clean(g(16)),
                "needs_review": False,
            }
    return rows, program_names


# =============================================================================
# ৪. Borrower List PDF-এর ভাঙা রো মেরামত (সহজ + জটিল মাল্টিলাইন কেস)
# =============================================================================
# Union-এর নাম গুলো সাধারণত পুনরাবৃত্ত ছোট সেট -- ভালো রো থেকে স্বয়ংক্রিয়ভাবে শেখা হবে
def build_known_unions(clean_rows):
    unions = set()
    for r in clean_rows.values():
        u = clean(r.get("union", ""))
        if u:
            unions.add(u)
    return unions


_BORR_SIMPLE_RE = re.compile(
    r'^(\d+)\s+(\d+(?:\.\d+)?)\s+(.+?)\s+([+\d][\d+,\s]{5,}\d)\s+'
    r'(\d{2}/\d{2}/\d{4})\s+(.+)$'
)


def _split_phone_ward(phone_blob):
    """'5 8801729547762' এর মতো ব্লব থেকে Ward No (যদি থাকে) আর আসল ফোন আলাদা করে"""
    parts = clean(phone_blob).split()
    if len(parts) >= 2 and len(parts[-1]) >= 8:
        return parts[-1]
    return "".join(parts)


def repair_borrower_row(merged_text, known_unions):
    """
    রিটার্ন করে dict:
      সবসময় নির্ভরযোগ্য: loan_case, phone, overdue_date, installment,
                          received_amount, current_balance, due_amount, reschedule_no
      best-effort:        village, union (known_unions সেট মিললে)
      needs_review=True হলে father/spouse/village আলাদা করা যায়নি -- raw_middle রাখা হবে
    """
    s = clean(merged_text)
    has_multiline_jumble = "\n" in str(merged_text)

    m = _BORR_SIMPLE_RE.match(s)
    if not m:
        return None

    sl, case, middle, phone_blob, odate, rest = m.groups()
    phone = _split_phone_ward(phone_blob)

    # rest = Installment ReceivedAmt CurrBal DueNo DueAmt RescheduleNo [RescheduleDate LastRecvAmt LastRecvDate NotPaid...]
    rtoks = rest.split()
    numeric_tail = {}
    # প্রথম ৩টা সবসময় থাকে: Installment, ReceivedAmount(often blank->skip), CurrentBalance
    # নির্ভরযোগ্যভাবে শুধু installment (প্রথম সংখ্যা) আর current_balance বের করা কঠিন কারণ
    # ReceivedAmount প্রায়ই blank থাকে -- তাই conservative approach: প্রথম টোকেন = installment
    installment = rtoks[0] if rtoks else ""

    # due amount: pattern "<no> <amount>" বা শুধু ফাঁকা -- খুঁজি একটা float-like no (যেমন '1.08')
    due_amount = ""
    for tok in rtoks:
        if re.match(r"^\d+\.\d+$", tok):
            idx = rtoks.index(tok)
            if idx + 1 < len(rtoks):
                due_amount = rtoks[idx + 1]
            break

    # reschedule no: single digit that stands alone near a date pattern in rest
    resched_no = ""
    date_positions = [i for i, t in enumerate(rtoks) if re.match(r"^\d{2}/\d{2}/\d{4}$", t)]
    if date_positions:
        first_date_idx = date_positions[0]
        if first_date_idx - 1 >= 0 and re.match(r"^\d+$", rtoks[first_date_idx - 1]):
            resched_no = rtoks[first_date_idx - 1]

    result = {
        "loan_case": case,
        "phone": phone,
        "overdue_date": odate,
        "installment": installment,
        "due_amount": due_amount,
        "reschedule_no": resched_no,
        "father": "",
        "spouse": "",
        "village": "",
        "union": "",
        "needs_review": True,
        "raw_middle": middle,
    }

    # Union known-set দিয়ে মেলানোর চেষ্টা (best-effort)
    for u in sorted(known_unions, key=len, reverse=True):
        if middle.strip().endswith(u):
            result["union"] = u
            remainder = middle[: -len(u)].strip()
            # remainder-এর শেষ শব্দ(গুলো) সাধারণত Village -- একটা শব্দ ধরে নিচ্ছি (best effort)
            rparts = remainder.rsplit(" ", 1)
            if len(rparts) == 2:
                result["village"] = rparts[1]
                result["raw_middle"] = rparts[0]  # বাকিটা Borrower+Father+Spouse -- review দরকার
            break

    if has_multiline_jumble:
        result["needs_review"] = True

    return result


def parse_borrower_list_pdf(pdf_path, progress_cb=None):
    """
    দুই-পাস:
      pass 1: সব "স্বাভাবিক" (ভাঙা নয়) রো পার্স করে known_unions শিখে নেয়
      pass 2: ভাঙা রো মেরামত করে (known_unions ব্যবহার করে)
    রিটার্ন করে: rows dict {(code, loan_case): {...}}, program_names {code: name}
    """
    sections, program_names = get_sections_and_tables(pdf_path, progress_cb)

    rows = {}
    broken_rows_raw = []  # (code, prog_code, prog_name, stype, raw_text)

    for prog_code, prog_name, stype, code, table in sections:
        for row in table:
            if not row:
                continue

            if row[0] and (len(row) < 2 or row[1] is None) and len(str(row[0]).split()) > 5:
                broken_rows_raw.append((code, prog_code, prog_name, stype, row[0]))
                continue

            if not row[1] or not re.match(r"^\d+(\.\d+)?$", str(row[1]).strip()):
                continue

            loan_case = normalize_case_key(row[1])
            key = (code, loan_case)

            def g(i):
                return row[i] if i < len(row) else None

            rows[key] = {
                "account_code": code,
                "program_code": prog_code,
                "program_name": prog_name,
                "status_type": stype,
                "loan_case": loan_case,
                "borrower": clean(g(2)),
                "father": clean(g(3)),
                "spouse": clean(g(4)),
                "village": clean(g(5)),
                "union": clean(g(7)),
                "phone": clean(g(9)),
                "overdue_date": clean(g(10)),
                "installment": clean(g(11)),
                "received_amount": clean(g(12)),
                "current_balance": clean(g(13)),
                "due_amount": clean(g(15)),
                "reschedule_no": clean(g(16)),
                "needs_review": False,
            }

    known_unions = build_known_unions(rows)

    for code, prog_code, prog_name, stype, raw_text in broken_rows_raw:
        repaired = repair_borrower_row(raw_text, known_unions)
        if not repaired:
            continue
        norm_case = normalize_case_key(repaired["loan_case"])
        key = (code, norm_case)
        rows[key] = {
            "account_code": code,
            "program_code": prog_code,
            "program_name": prog_name,
            "status_type": stype,
            "loan_case": norm_case,
            "borrower": "",  # loan_balance PDF থেকে merge()-এর সময় বসবে
            "father": repaired["father"],
            "spouse": repaired["spouse"],
            "village": repaired["village"],
            "union": repaired["union"],
            "phone": repaired["phone"],
            "overdue_date": repaired["overdue_date"],
            "installment": repaired["installment"],
            "received_amount": "",
            "current_balance": "",
            "due_amount": repaired["due_amount"],
            "reschedule_no": repaired["reschedule_no"],
            "needs_review": True,
            "raw_middle": repaired.get("raw_middle", ""),
        }

    return rows, program_names


# =============================================================================
# ৫. দুই সোর্স মেলানো (নাম-মিল যাচাইসহ)
# =============================================================================
def name_similarity(a, b):
    a, b = clean(a).lower(), clean(b).lower()
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def merge_sources(borrower_rows, balance_rows, prefix_map, name_match_threshold=0.55):
    merged = []
    all_keys = set(borrower_rows) | set(balance_rows)

    for key in all_keys:
        code, loan_case = key
        b = borrower_rows.get(key, {})
        bal = balance_rows.get(key, {})

        borrower_name = b.get("borrower") or bal.get("customer_name") or ""
        bal_name = bal.get("customer_name", "")

        name_ok = True
        if borrower_name and bal_name:
            name_ok = name_similarity(borrower_name, bal_name) >= name_match_threshold

        # prefix সবসময় Loan Program কোড (যেমন '0101') দিয়ে বসে -- account_code ('100101') দিয়ে নয়
        program_code = b.get("program_code") or bal.get("program_code") or ""
        prefix = prefix_map.get(program_code, program_code or code)
        needs_review = bool(b.get("needs_review")) or bool(bal.get("needs_review")) or (not name_ok)

        note_parts = []
        if key not in borrower_rows:
            note_parts.append("Borrower List-এ পাওয়া যায়নি")
        if key not in balance_rows:
            note_parts.append("Loan Balance-এ পাওয়া যায়নি")
        if not name_ok:
            note_parts.append(f"নাম অমিল সন্দেহ ({borrower_name!r} vs {bal_name!r})")
        if b.get("needs_review"):
            note_parts.append("Father/Spouse/Village wrap সমস্যা - ম্যানুয়াল চেক দরকার")

        merged.append({
            "prefixed_loan_case": f"{prefix}-{loan_case}",
            "account_code": code,
            "loan_case_no": loan_case,
            "borrower": borrower_name,
            "father": b.get("father", ""),
            "spouse": b.get("spouse", ""),
            "village": b.get("village", ""),
            "union": b.get("union", ""),
            "phone": bal.get("phone", "") or b.get("phone", ""),
            "overdue_date": bal.get("overdue_date", "") or b.get("overdue_date", ""),
            "installment": b.get("installment", ""),
            "bal_principal": bal.get("bal_principal", ""),
            "bal_interest": bal.get("bal_interest", ""),
            "bal_total": bal.get("bal_total", ""),
            "due_amount": b.get("due_amount", ""),
            "reschedule_no": b.get("reschedule_no", ""),
            "needs_review": needs_review,
            "review_note": " | ".join(note_parts),
        })
    return merged


# =============================================================================
# ৬. Excel এ লেখা
# =============================================================================
def _sort_key(d):
    try:
        case_num = float(re.sub(r"[^\d.]", "", d["loan_case_no"]) or 0)
    except ValueError:
        case_num = 0
    return (d["account_code"], case_num)


def write_excel(rows, out_path):
    rows = sorted(rows, key=_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"

    # দুই-স্তরের হেডার (Balance-এর জন্য sub-columns)
    top_headers = ["Prefixed Loan Case", "Borrower", "Father", "Spouse", "Village",
                   "Union", "Phone", "Overdue Date", "Installment Amount",
                   "Balance", "", "", "Due Amount", "Reschedule No.", "যাচাই দরকার?", "নোট"]
    sub_headers = ["", "", "", "", "", "", "", "", "",
                   "Principal", "Interest", "Total", "", "", "", ""]

    ws.append(top_headers)
    ws.append(sub_headers)
    ws.merge_cells("J1:L1")
    for col in list("ABCDEFGHI") + ["M", "N", "O", "P"]:
        ws.merge_cells(f"{col}1:{col}2")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for row_cells in ws.iter_rows(min_row=1, max_row=2):
        for c in row_cells:
            c.font = Font(name="Arial", bold=True)
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for d in rows:
        ws.append([
            d["prefixed_loan_case"], d["borrower"], d["father"], d["spouse"],
            d["village"], d["union"], d["phone"], d["overdue_date"],
            to_num(d["installment"]),
            to_num(d["bal_principal"]), to_num(d["bal_interest"]), to_num(d["bal_total"]),
            to_num(d["due_amount"]), d["reschedule_no"] or None,
            "হ্যাঁ" if d["needs_review"] else "",
            d["review_note"],
        ])

    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows(min_row=3):
        for c in row:
            c.font = Font(name="Arial")
        if row[14].value == "হ্যাঁ":
            for c in row:
                c.fill = review_fill

    widths = [16, 20, 20, 16, 14, 12, 14, 12, 12, 12, 12, 12, 11, 11, 10, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    wb.save(out_path)

    total = len(rows)
    review = sum(1 for d in rows if d["needs_review"])
    return {"total_rows": total, "needs_review": review}


# =============================================================================
# ৭. মূল প্রসেসিং ফাংশন (app.py থেকে কল হবে)
# =============================================================================
def process(borrower_pdf_path, balance_pdf_path, out_xlsx_path,
            prefix_overrides=None, progress_cb=None):
    """
    প্রধান এন্ট্রি পয়েন্ট। রিটার্ন করে সামারি dict।
    """
    borrower_rows, borrower_programs = parse_borrower_list_pdf(
        borrower_pdf_path, progress_cb=lambda p, t: progress_cb("borrower", p, t) if progress_cb else None
    )
    balance_rows, balance_programs = parse_loan_balance_pdf(
        balance_pdf_path, progress_cb=lambda p, t: progress_cb("balance", p, t) if progress_cb else None
    )

    all_programs = {**balance_programs, **borrower_programs}
    prefix_map = build_prefix_map(all_programs, prefix_overrides)

    merged = merge_sources(borrower_rows, balance_rows, prefix_map)
    stats = write_excel(merged, out_xlsx_path)
    stats["programs"] = all_programs
    stats["prefix_map"] = prefix_map
    stats["output_path"] = out_xlsx_path
    return stats
